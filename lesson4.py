#!/usr/bin/env python
#!-*- coding:utf-8 -*-
#@Time   :2020/7/1517:48
#@Email  :hercules583@163.com
#File    :lesson4.py
#software:PyCharm
import requests
import openpyxl
def read_data(filename,sheetname):
    wb=openpyxl.load_workbook(filename)#加载工作簿
    sheet=wb[sheetname]#获取表单
    max_raw=sheet.max_row#获取最大行数
    case_list=[]#创建空列表，存放测试用例
    for i in range(2,max_raw+1):
        dict1=dict(
        case_id=sheet.cell(row=i,column=1).value,#获取case——id
        url=sheet.cell(row=i,column=5).value,#获取url
        data=sheet.cell(row=i,column=6).value,#获取data传入的参数
        expect=sheet.cell(row=i,column=7).value#获取期望结果
        )
        case_list.append(dict1)#每循环一次就将一行数据加入列表
    return case_list#返回测试用例列表

def register(url,data):
    headers_reg = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    re=requests.post(url=url,json=data,headers=headers_reg)
    response=re.json()
    return response

def write_result(filename,sheetname,row,column,final_result):
    wb=openpyxl.load_workbook(filename)#加载文件
    sheet=wb[sheetname]#读取sheet
    sheet.cell(row=row,column=column).value=final_result#将结果写入cell
    wb.save(filename)#保存excel

cases=read_data("test_case_api.xlsx","register")#调用读取excel文件函数
for case in cases:
    case_id=case.get("case_id")#从列表中取出case_id
    url=case.get("url")#从列表中取出url
    data=eval(case.get("data"))#从列表中取出data
    expect=eval(case.get("expect"))#从列表中取出expect
    expect_msg=expect.get("msg")#从except取出msg
    real_result=register(url=url,data=data)#调用注册接口函数将从列表取出的url，data作为参数注册函数
    real_msg=real_result.get("msg")
    # print("预期结果的msg:{}".format(expect_msg))
    # print("实际结果的msg:{}".format(real_msg))
    if real_msg==expect_msg:
        print("第{}条测试用例通过".format(case_id))
        final_re="passed"
    else:
        print("第{}条测试用例不通过".format(case_id))
        final_re="failed"
    write_result("test_case_api.xlsx","register",case_id+1,8,final_re)#调用写入函数

